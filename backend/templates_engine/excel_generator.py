"""
backend/templates_engine/excel_generator.py
============================================
Generates a formatted, multi-sheet Excel report (.xlsx) using OpenPyXL.

Public API
----------
generate_excel_report(report_type, summary_data, df_data) -> io.BytesIO

Parameters
----------
report_type  : str             "sales" | "user_activity" | "inventory"
summary_data : dict            Output of get_*_summary() – may contain scalar
                               values AND pandas DataFrames for breakdown tables.
df_data      : pd.DataFrame    Raw detail rows from get_*_data().

Returns
-------
io.BytesIO  with seek position at 0, ready to stream or save.
"""

import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── colour palette ──────────────────────────────────────────────────────────
_DARK_BLUE   = "1F4E78"
_MID_BLUE    = "2C3E50"
_LIGHT_BLUE  = "D9E1F2"
_ALT_ROW     = "EEF2F8"
_WHITE       = "FFFFFF"
_BLACK       = "000000"
_ACCENT      = "2980B9"


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int,
                     header_fill_hex: str = _MID_BLUE) -> int:
    """
    Write a DataFrame into *ws* starting at *start_row*.
    Returns the next available row after the table.
    """
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="No data available.").font = Font(italic=True, color="888888")
        return start_row + 2

    header_fill = PatternFill(start_color=header_fill_hex,
                              end_color=header_fill_hex, fill_type="solid")
    alt_fill    = PatternFill(start_color=_ALT_ROW, end_color=_ALT_ROW, fill_type="solid")
    border      = _thin_border()

    # Header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx,
                       value=col_name.replace("_", " ").title())
        cell.font      = Font(name="Arial", size=10, bold=True, color=_WHITE)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    # Data rows
    for r_offset, (_, row) in enumerate(df.iterrows(), 1):
        fill = alt_fill if r_offset % 2 == 0 else None
        for col_idx, value in enumerate(row, 1):
            # Resolve pandas Timestamps / numpy types
            if hasattr(value, "item"):          # numpy scalar → Python native
                value = value.item()
            if hasattr(value, "date"):          # Timestamp → date string
                value = value.strftime("%Y-%m-%d")

            cell = ws.cell(row=start_row + r_offset, column=col_idx, value=value)
            cell.font   = Font(name="Arial", size=9)
            cell.border = border
            if fill:
                cell.fill = fill
            if isinstance(value, float):
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")
            elif isinstance(value, int):
                cell.number_format = "#,##0"
                cell.alignment     = Alignment(horizontal="right")

    # Auto-fit columns
    for col in ws.iter_cols(min_row=start_row,
                             max_row=start_row + len(df),
                             min_col=1, max_col=len(df.columns)):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 12), 45)

    return start_row + len(df) + 2   # +1 for header, +1 gap


def generate_excel_report(report_type: str,
                           summary_data: dict,
                           df_data: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()

    # ── Sheet 1 – Summary ────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    title_font  = Font(name="Arial", size=18, bold=True, color=_WHITE)
    bold_font   = Font(name="Arial", size=11, bold=True)
    label_font  = Font(name="Arial", size=10)
    kpi_font    = Font(name="Arial", size=11, bold=True, color=_DARK_BLUE)
    sub_font    = Font(name="Arial", size=12, bold=True, color=_MID_BLUE)

    title_fill = PatternFill(start_color=_DARK_BLUE, end_color=_DARK_BLUE, fill_type="solid")
    kpi_fill   = PatternFill(start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid")

    # Title banner (rows 1–2)
    report_label = report_type.replace("_", " ").title()
    ws_summary.merge_cells("A1:H2")
    title_cell = ws_summary["A1"]
    title_cell.value     = f"CUSTOM REPORT  ·  {report_label.upper()}  ·  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font      = title_font
    title_cell.fill      = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 20
    ws_summary.row_dimensions[2].height = 20

    current_row = 4

    # ── KPI scalars (non-DataFrame values in summary_data) ──────────────────
    scalar_items = {k: v for k, v in summary_data.items() if not isinstance(v, pd.DataFrame)}
    if scalar_items:
        ws_summary.cell(row=current_row, column=1, value="Key Performance Indicators").font = sub_font
        current_row += 1

        for key, value in scalar_items.items():
            label_cell = ws_summary.cell(row=current_row, column=1,
                                         value=key.replace("_", " ").title())
            label_cell.font   = label_font
            label_cell.fill   = kpi_fill
            label_cell.border = _thin_border()

            val_cell = ws_summary.cell(row=current_row, column=2, value=value)
            val_cell.font   = kpi_font
            val_cell.fill   = kpi_fill
            val_cell.border = _thin_border()
            if isinstance(value, float):
                val_cell.number_format = "#,##0.00"
                val_cell.alignment     = Alignment(horizontal="right")
            elif isinstance(value, int):
                val_cell.number_format = "#,##0"
                val_cell.alignment     = Alignment(horizontal="right")
            current_row += 1

        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 20
        current_row += 2

    # ── Breakdown DataFrames in summary ─────────────────────────────────────
    df_items = {k: v for k, v in summary_data.items() if isinstance(v, pd.DataFrame)}
    for section_name, section_df in df_items.items():
        ws_summary.cell(row=current_row, column=1,
                        value=section_name.replace("_", " ").title()).font = sub_font
        current_row += 1
        current_row = _write_dataframe(ws_summary, section_df, current_row)

    # ── Sheet 2 – Detail Records ─────────────────────────────────────────────
    ws_detail = wb.create_sheet(title="Detail Records")
    ws_detail.merge_cells("A1:J1")
    detail_title = ws_detail["A1"]
    detail_title.value     = f"{report_label} – Detailed Records"
    detail_title.font      = Font(name="Arial", size=14, bold=True, color=_WHITE)
    detail_title.fill      = PatternFill(start_color=_ACCENT, end_color=_ACCENT, fill_type="solid")
    detail_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_detail.row_dimensions[1].height = 22

    _write_dataframe(ws_detail, df_data, start_row=3, header_fill_hex=_ACCENT)

    # ── Save ──────────────────────────────────────────────────────────────────
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
